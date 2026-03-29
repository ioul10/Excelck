"""FiscalXL v5 — Un seul parseur X/Y (pdfplumber) pour DGI et AMMC"""

import streamlit as st
import tempfile, os
from pathlib import Path
import pandas as pd
import openpyxl

from core.pdf_parser import PDFParser
from core.injector   import TemplateInjector
from utils.validator import validate_pdf_structure_v2
from utils.logger    import get_logger

logger = get_logger(__name__)
TEMPLATE_PATH = Path(__file__).parent / "template_fiscal.xlsx"

st.set_page_config(page_title="FiscalXL v5", page_icon="📊", layout="wide",
                   initial_sidebar_state="expanded")

st.markdown("""
<style>
.main-header{background:linear-gradient(135deg,#1F3864,#2E75B6);padding:1.5rem 2rem;
  border-radius:12px;margin-bottom:1.2rem;}
.main-header h1{color:white;margin:0;font-size:1.8rem;}
.main-header p{color:#BDD7EE;margin:.3rem 0 0;font-size:.88rem;}
.kpi-box{background:white;border:1px solid #BDD7EE;border-radius:8px;
  padding:.7rem;text-align:center;}
.kpi-box .val{font-size:1.15rem;font-weight:bold;color:#1F3864;}
.kpi-box .lbl{font-size:.72rem;color:#888;margin-top:.2rem;}
.success-box{background:#E2EFDA;border:1px solid #70AD47;border-radius:8px;
  padding:.9rem 1.3rem;color:#375623;}
.warn-box{background:#FFF2CC;border:1px solid #FFD700;border-radius:8px;
  padding:.7rem 1.1rem;color:#7B5900;}
.error-box{background:#FCE4D6;border:1px solid #C55A11;border-radius:8px;
  padding:.9rem 1.3rem;color:#7B2C00;}
.badge-dgi{background:#1F3864;color:white;padding:3px 10px;
  border-radius:20px;font-size:.78rem;font-weight:bold;}
.badge-ammc{background:#70AD47;color:white;padding:3px 10px;
  border-radius:20px;font-size:.78rem;font-weight:bold;}
div[data-testid="stDownloadButton"] button{
  background:linear-gradient(135deg,#1F3864,#2E75B6);color:white;
  border:none;padding:.8rem 2.5rem;font-size:1rem;border-radius:8px;width:100%;}
</style>""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
  <h1>📊 FiscalXL v5</h1>
  <p>Convertisseur PDF → Excel · Pièces annexes IS (Modèle Comptable Normal, loi 9-88 Maroc)</p>
</div>""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Configuration")
    mode = st.radio(
        "**Format PDF**",
        options=["🏛️ DGI — 7 pages", "📄 AMMC / Standard — 5 pages"],
        index=0,
        help=(
            "**DGI** : Étatde synthèse DGI (7 pages)\n\n"
            "**AMMC** : Liasse standard (5 pages)"
        ),
    )
    is_dgi = "DGI" in mode

    st.markdown("---")
    show_preview    = st.toggle("Aperçu des données extraites", value=True)
    show_comparison = st.toggle("Comparaison PDF ↔ Excel",      value=True)
    show_debug      = st.toggle("Mode débogage",                value=False)
    st.markdown("---")

    if is_dgi:
        st.markdown('<span class="badge-dgi">MODE DGI — 7 PAGES</span>',
                    unsafe_allow_html=True)
        st.markdown("""
        - **Page 1** — Identification
        - **Pages 2-3** — Bilan Actif
        - **Page 4** — Bilan Passif
        - **Pages 5-7** — CPC
        """)
    else:
        st.markdown('<span class="badge-ammc">MODE AMMC — 5 PAGES</span>',
                    unsafe_allow_html=True)
        st.markdown("""
        - **Page 1** — Identification
        - **Page 2** — Bilan Actif
        - **Page 3** — Bilan Passif
        - **Pages 4-5** — CPC
        """)

    st.success("✅ Template chargé") if TEMPLATE_PATH.exists() else st.error("⚠️ Template manquant")
    st.caption("FiscalXL v5 · MCN loi 9-88")

# ── Helpers ───────────────────────────────────────────────────────────────────

def update_excel_headers(wb, info: dict):
    raison   = info.get("raison_sociale") or "—"
    id_fisc  = info.get("identifiant_fiscal") or ""
    exercice = info.get("exercice") or ""
    sub = f"{raison}  —  IF: {id_fisc}" if id_fisc else raison

    for sheet, title in [
        ("2 - Bilan Actif",   f"BILAN — ACTIF  |  {exercice}"),
        ("3 - Bilan Passif",  f"BILAN — PASSIF  |  {exercice}"),
        ("4 - CPC",           f"COMPTE DE PRODUITS ET CHARGES  |  {exercice}"),
        ("5 - Tableau de Bord","TABLEAU DE BORD — SYNTHÈSE FINANCIÈRE"),
    ]:
        if sheet not in wb.sheetnames: continue
        wb[sheet]["A1"] = title
        wb[sheet]["A2"] = sub if "Bord" not in sheet else raison

    if "1 - Infos Générales" in wb.sheetnames:
        ws = wb["1 - Infos Générales"]
        mapping = {
            "Raison sociale":       raison,
            "Identifiant fiscal":   id_fisc,
            "Taxe professionnelle": info.get("taxe_professionnelle") or "—",
            "Adresse":              info.get("adresse") or "—",
            "Exercice":             exercice,
            "Date de déclaration":  info.get("date_declaration") or "—",
            "Nombre de pages":      str(info.get("pages") or "—"),
        }
        for row in ws.iter_rows():
            if row[0].value in mapping:
                row[1].value = mapping[row[0].value]


def build_comparison(extracted: dict, wb) -> list:
    checks = [
        ("actif",  "Terrains",                        "2 - Bilan Actif",  "B15","E15","Terrains"),
        ("actif",  "Constructions",                   "2 - Bilan Actif",  "B16","E16","Constructions"),
        ("actif",  "Installations techniques",        "2 - Bilan Actif",  "B17","E17","Installations techniques"),
        ("actif",  "Clients et comptes rattachés",    "2 - Bilan Actif",  "B40","E40","Clients et ctes rattachés"),
        ("actif",  "Banques",                         "2 - Bilan Actif",  "B51","E51","Banques T.G et C.C.P"),
        ("passif", "Capital social ou personnel",     "3 - Bilan Passif", "B6", "C6", "Capital social"),
        ("passif", "Résultat net de l'exercice",      "3 - Bilan Passif", "B15","C15","Résultat net exercice"),
        ("passif", "Autres dettes de financement",    "3 - Bilan Passif", "B22","C22","Autres dettes financement"),
        ("passif", "Fournisseurs et comptes rattachés","3 - Bilan Passif","B30","C30","Fournisseurs"),
        ("cpc",    "Ventes de biens et services",     "4 - CPC",          "B6", "E6", "Ventes biens/services"),
        ("cpc",    "Achats consommés",                "4 - CPC",          "B11","E11","Achats consommés"),
        ("cpc",    "Charges de personnel",            "4 - CPC",          "B19","E19","Charges de personnel"),
        ("cpc",    "Dotations d'exploitation",        "4 - CPC",          "B20","E20","Dotations d'exploitation"),
        ("cpc",    "IMPOTS SUR LES",                  "4 - CPC",          "B53","E53","Impôts sur résultats"),
    ]

    src = {
        "actif":  extracted.get("actif_values",  {}),
        "passif": extracted.get("passif_values", {}),
        "cpc":    extracted.get("cpc_values",    {}),
    }

    rows = []
    for section, skey, sheet, cn, cn1, display in checks:
        pdf_v = next(
            (v for k, v in src[section].items()
             if skey.lower() in k.lower() or k.lower() in skey.lower()),
            None
        )
        pdf_n  = pdf_v[0] if pdf_v else None
        pdf_n1 = (pdf_v[2] if section == "actif" and pdf_v and len(pdf_v) > 2
                  else (pdf_v[1] if pdf_v and len(pdf_v) > 1 else None))

        xl_n = xl_n1 = None
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            for cell, target in [(cn, "n"), (cn1, "n1")]:
                try:
                    v = ws[cell].value
                    if isinstance(v, (int, float)):
                        if target == "n":  xl_n  = float(v)
                        else:              xl_n1 = float(v)
                except Exception: pass

        def fmt(v): return f"{v:,.2f}" if v is not None else "—"
        def st_s(a, b):
            if a is None and b is None: return "vide"
            if a is None or b is None:  return "manquant"
            return "ok" if abs(a - b) < 1 else "erreur"

        rows.append({
            "Section":  section.upper(), "Poste": display,
            "Cel.N": cn,   "PDF N": fmt(pdf_n),   "Excel N": fmt(xl_n),   "✓N":  st_s(pdf_n,  xl_n),
            "Cel.N-1":cn1, "PDF N-1":fmt(pdf_n1), "Excel N-1":fmt(xl_n1), "✓N-1":st_s(pdf_n1, xl_n1),
        })
    return rows


def render_comparison(rows: list):
    errors   = [r for r in rows if "erreur"   in (r["✓N"], r["✓N-1"])]
    warnings = [r for r in rows if "manquant" in (r["✓N"], r["✓N-1"])]
    ok_count = sum(1 for r in rows if r["✓N"] == "ok")

    c1, c2, c3 = st.columns(3)
    c1.metric("✅ Postes corrects",    ok_count)
    c2.metric("⚠️ Valeurs manquantes", len(warnings))
    c3.metric("❌ Erreurs de valeur",   len(errors))

    if errors:
        st.markdown("#### ❌ Erreurs — à corriger manuellement dans l'Excel")
        cols = ["Section","Poste","Cel.N","PDF N","Excel N","Cel.N-1","PDF N-1","Excel N-1"]
        st.dataframe(pd.DataFrame(errors)[cols], use_container_width=True,
                     height=min(220, 50+len(errors)*38))

    if warnings:
        st.markdown("#### ⚠️ Postes non injectés (valeur absente dans le PDF)")
        cols = ["Section","Poste","Cel.N","PDF N","Excel N"]
        st.dataframe(pd.DataFrame(warnings)[cols], use_container_width=True,
                     height=min(180, 50+len(warnings)*38))

    if not errors and not warnings:
        st.success("✅ Tous les postes vérifiés sont corrects !")

    with st.expander("📋 Tableau complet de comparaison"):
        badge = {"ok":"✅","manquant":"⚠️","erreur":"❌","vide":"—"}
        disp = [{
            "Section":r["Section"],"Poste":r["Poste"],
            "Cellule N":r["Cel.N"],"PDF N":r["PDF N"],"Excel N":r["Excel N"],"":badge.get(r["✓N"],"?"),
            "Cellule N-1":r["Cel.N-1"],"PDF N-1":r["PDF N-1"],"Excel N-1":r["Excel N-1"]," ":badge.get(r["✓N-1"],"?"),
        } for r in rows]
        st.dataframe(pd.DataFrame(disp), use_container_width=True, height=520)


# ── Layout principal ─────────────────────────────────────────────────────────
col_up, col_pipe = st.columns([3, 2])
with col_up:
    st.markdown("### 📂 Importer le PDF")
    uploaded = st.file_uploader("Glissez-déposez ou cliquez", type=["pdf"])

with col_pipe:
    st.markdown("### 🔄 Pipeline")
    for step, desc in [
        ("1 · Lecture PDF",     "pdfplumber — coordonnées X/Y"),
        ("2 · Extraction",      "Labels + valeurs par colonnes"),
        ("3 · Headers",         "Raison sociale et exercice dynamiques"),
        ("4 · Injection Excel", "Valeurs → cellules du template"),
        ("5 · Comparaison",     "PDF ↔ Excel poste par poste"),
    ]:
        color = "#1F3864" if is_dgi else "#70AD47"
        st.markdown(
            f'<div style="background:#f8f9fa;border-left:4px solid {color};'
            f'padding:.5rem .8rem;border-radius:0 6px 6px 0;margin:.3rem 0;">'
            f'<strong style="color:#1F3864">{step}</strong><br>'
            f'<span style="color:#555;font-size:.83rem">{desc}</span></div>',
            unsafe_allow_html=True)

if not uploaded:
    st.markdown("""
    <div style="text-align:center;padding:3rem;color:#888;border:2px dashed #BDD7EE;
      border-radius:12px;background:#f8fafd;margin-top:1rem;">
      <div style="font-size:3rem;">📄</div>
      <h3 style="color:#2E75B6;">Importez un PDF pour commencer</h3>
      <p>Sélectionnez le format dans la sidebar, puis importez votre fichier</p>
      <p style="font-size:.85rem;">
        <strong>DGI (7 pages)</strong> &nbsp;·&nbsp; <strong>AMMC (5 pages)</strong>
      </p>
    </div>""", unsafe_allow_html=True)
    st.stop()

# ── Processing ────────────────────────────────────────────────────────────────
st.markdown("---")
if not TEMPLATE_PATH.exists():
    st.markdown('<div class="error-box">❌ Template manquant.</div>', unsafe_allow_html=True)
    st.stop()

with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
    tmp.write(uploaded.getbuffer())
    pdf_path = tmp.name
output_path = pdf_path.replace(".pdf", "_out.xlsx")

try:
    progress = st.progress(0)
    status   = st.empty()

    # Étape 1 — Validation
    status.info("🔍 Étape 1/5 — Validation du PDF...")
    progress.progress(8)
    parser     = PDFParser(pdf_path)
    validation = validate_pdf_structure_v2(parser, mode="dgi" if is_dgi else "ammc")

    if not validation["valid"]:
        st.markdown(f'<div class="error-box">⚠️ {validation["message"]}</div>',
                    unsafe_allow_html=True)
        st.stop()

    meta    = validation["meta"]
    n_pages = meta.get("pages", 0)

    if is_dgi and n_pages != 7:
        st.warning(f"⚠️ Mode DGI sélectionné mais le PDF a {n_pages} pages (attendu : 7).")
    elif not is_dgi and n_pages not in (5, 6):
        st.warning(f"⚠️ Mode AMMC sélectionné mais le PDF a {n_pages} pages (attendu : 5).")

    badge_html = (f'<span class="badge-dgi">DGI</span>' if is_dgi
                  else f'<span class="badge-ammc">AMMC</span>')

    for col, (lbl, val) in zip(
        st.columns(5),
        [("Raison Sociale",    (meta.get("raison_sociale") or "—")[:20]),
         ("Identifiant Fiscal", meta.get("identifiant_fiscal") or "—"),
         ("Fin exercice",       meta.get("exercice_fin") or "—"),
         ("Pages",              str(n_pages)),
         ("Format",             "DGI — 7p" if is_dgi else "AMMC — 5p")]
    ):
        col.markdown(f'<div class="kpi-box"><div class="val">{val}</div>'
                     f'<div class="lbl">{lbl}</div></div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    progress.progress(20)

    # Étape 2 — Extraction
    status.info("📄 Étape 2/5 — Extraction des valeurs...")
    progress.progress(38)
    with st.spinner("Parsing en cours..."):
        extracted = parser.parse()
    progress.progress(58)

    if show_debug:
        with st.expander("🐛 Données brutes"):
            st.json({k: {str(l): v for l, v in d.items()} if isinstance(d, dict) else d
                     for k, d in extracted.items()})

    # Étape 3 — Injection
    status.info("🔗 Étape 3/5 — Injection dans le template...")
    progress.progress(65)
    with st.spinner("Injection..."):
        inj   = TemplateInjector(str(TEMPLATE_PATH))
        stats = inj.inject(extracted, output_path)
    progress.progress(78)

    # Étape 4 — Headers dynamiques
    status.info("🏷️ Étape 4/5 — Mise à jour des headers...")
    wb = openpyxl.load_workbook(output_path)
    update_excel_headers(wb, extracted["info"])
    wb.save(output_path)
    progress.progress(88)

    # Étape 5 — Vérification
    status.info("✅ Étape 5/5 — Vérification...")
    wb_check = openpyxl.load_workbook(output_path)
    n_formulas = sum(
        1 for ws in wb_check.worksheets
        for row in ws.iter_rows()
        for c in row if isinstance(c.value, str) and c.value.startswith("=")
    )
    progress.progress(100)
    status.empty()

    # Résumé
    st.markdown(f"""
    <div class="success-box">
      ✅ <strong>Fichier Excel généré !</strong>
      &nbsp;·&nbsp; {len(wb_check.sheetnames)} feuilles
      &nbsp;·&nbsp; {n_formulas} formules intactes
      &nbsp;·&nbsp; {stats['injected']} valeurs injectées
      &nbsp;·&nbsp; {badge_html}
    </div>""", unsafe_allow_html=True)

    if stats.get("not_found"):
        nf = stats["not_found"]
        real = [x for x in nf if len(x) > 4 and not any(
            t in x.upper() for t in ["TOTAL", "CAPITAUX PROPRES", "RESULTAT D'EX",
                                      "CHARGES D'EX", "PRODUITS D'EX"]
        )]
        if real:
            st.markdown(
                f'<div class="warn-box">⚠️ <strong>{len(real)} postes non mappés</strong>'
                f' — valeurs extraites mais cellule non trouvée :<br>'
                f'<code>{"  ·  ".join(real[:8])}</code></div>',
                unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Aperçu
    if show_preview:
        with st.expander("👁️ Aperçu des valeurs extraites", expanded=False):
            tabs = st.tabs(["ℹ️ Infos", "📋 Bilan Actif", "📋 Bilan Passif", "📈 CPC"])
            with tabs[0]:
                for k, v in extracted.get("info", {}).items():
                    if k != "pages" and v:
                        st.markdown(f"**{k.replace('_',' ').title()}** : {v}")
            with tabs[1]:
                av = extracted.get("actif_values", {})
                if av:
                    df = pd.DataFrame(
                        [(k,
                          f"{v[0]:,.2f}" if v and v[0] is not None else "—",
                          f"{v[1]:,.2f}" if len(v)>1 and v[1] is not None else "—",
                          f"{v[2]:,.2f}" if len(v)>2 and v[2] is not None else "—")
                         for k, v in av.items()],
                        columns=["Poste","Brut","Amort.","Net N-1"])
                    st.dataframe(df, use_container_width=True, height=320)
            with tabs[2]:
                pv = extracted.get("passif_values", {})
                if pv:
                    df = pd.DataFrame(
                        [(k,
                          f"{v[0]:,.2f}" if v and v[0] is not None else "—",
                          f"{v[1]:,.2f}" if len(v)>1 and v[1] is not None else "—")
                         for k, v in pv.items()],
                        columns=["Poste","Exercice N","Exercice N-1"])
                    st.dataframe(df, use_container_width=True, height=320)
            with tabs[3]:
                cv = extracted.get("cpc_values", {})
                if cv:
                    df = pd.DataFrame(
                        [(k,
                          f"{v[0]:,.2f}" if v and v[0] is not None else "—",
                          f"{v[1]:,.2f}" if len(v)>1 and v[1] is not None else "—",
                          f"{v[2]:,.2f}" if len(v)>2 and v[2] is not None else "—")
                         for k, v in cv.items()],
                        columns=["Désignation","Propre N","Exerc. Préc.","Total N-1"])
                    st.dataframe(df, use_container_width=True, height=320)

    # Comparaison
    if show_comparison:
        st.markdown("### 🔍 Comparaison PDF ↔ Excel")
        render_comparison(build_comparison(extracted, wb_check))

    # Téléchargement
    st.markdown("### ⬇️ Télécharger")
    raison_slug = (extracted["info"].get("raison_sociale") or "fiscal").replace(" ","_")[:20]
    date_slug   = (extracted["info"].get("exercice_fin") or "").replace("/","-")
    fname = f"{raison_slug}_{date_slug}_fiscal.xlsx"
    with open(output_path, "rb") as f:
        st.download_button("📥 Télécharger le fichier Excel", data=f, file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

except Exception as e:
    logger.exception("Erreur pipeline")
    st.markdown(f'<div class="error-box">❌ <strong>Erreur :</strong> <code>{e}</code></div>',
                unsafe_allow_html=True)
    if show_debug:
        import traceback; st.code(traceback.format_exc())
finally:
    for f in [pdf_path, output_path]:
        try:
            if os.path.exists(f): os.unlink(f)
        except Exception: pass
