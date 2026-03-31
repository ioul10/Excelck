"""
core/pdf_to_excel.py
Extraction PDF fiscal → Excel structuré.
Supporte AMMC (5 pages) et DGI (7 pages).
Méthode : extract_tables() avec fallback X/Y.
"""

import re, unicodedata
from collections import defaultdict
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.logger import get_logger

logger = get_logger(__name__)

# ── Couleurs ──────────────────────────────────────────────────────────────────
C_DARK   = "1F3864"
C_MED    = "2E75B6"
C_LIGHT  = "D6E4F0"
C_WHITE  = "FFFFFF"
C_GRAY   = "F5F7FA"
C_BORDER = "B8CCE4"
NUM_FMT  = '#,##0.00;[Red]-#,##0.00;"-"'

# ── Lignes à ignorer ──────────────────────────────────────────────────────────
SKIP_RE = re.compile(
    r'^(tableau|bilan\s*\(|compte\s*de\s*produits|agence\s*du|'
    r'identifiant\s*fiscal|exercice\s*du|fes\s*le|casablanca|'
    r'signature|cadre\s*reserve|modele\s*normal|nb\s*:|'
    r'\(1\)|\(2\)|\d+\s*=\s*\d+|propres\s*à|concernant\s*les|'
    r'totaux\s*de\s*l|brut\s*exercice|amortissements\s*et|'
    r'net\s*exercice|etat\s*de\s*synth)',
    re.IGNORECASE
)

ROTATED_RE = re.compile(r'^([A-Z]\n){3,}')  # lettres rotatives

def _should_skip(label: str) -> bool:
    if not label or len(label.strip()) < 2: return True
    s = label.strip()
    if ROTATED_RE.match(s.replace(' ', '\n')): return True
    # Séquences de lettres séparées par espaces (A C T I F)
    if re.match(r'^([A-Z]\s){3,}', s): return True
    if SKIP_RE.match(s): return True
    if re.match(r'^\d+$', s.strip()): return True
    return False

TOTAL_RE   = re.compile(r'^(total|sous-total)', re.I)
RESULT_RE  = re.compile(r'^(r[eé]sultat|impots?\s+sur|impôts?\s+sur)', re.I)
SECTION_RE = re.compile(r'^[A-ZÀÂÉÈÊÎÔÙÛÇ\s\(\)]{5,}$')

def _row_type(label: str) -> str:
    s = label.strip()
    if TOTAL_RE.match(s): return 'total'
    if RESULT_RE.match(s): return 'result'
    if SECTION_RE.match(s) and len(s) > 4: return 'section'
    return 'normal'


# ── Parsing numérique ─────────────────────────────────────────────────────────
def _parse_fr(s) -> float | None:
    if not s: return None
    s = str(s).strip().replace('\xa0', '').replace(' ', '')
    if not s or s in ['-', '—', '/', '']: return None
    neg = s.startswith('-'); s = s.lstrip('-')
    m = re.match(r'^(\d{1,3}(?:\.\d{3})*),(\d{2})$', s)
    if m: s = m.group(1).replace('.', '') + '.' + m.group(2)
    elif re.match(r'^\d+,\d{2}$', s): s = s.replace(',', '.')
    elif re.match(r'^\d+$', s): pass
    else: return None
    try: return -float(s) if neg else float(s)
    except: return None


# ── Extraction tableaux ───────────────────────────────────────────────────────
def _table_ok(table) -> bool:
    if not table or len(table) < 3: return False
    if len(table[0]) < 2: return False
    # Compter les lignes avec label séparé + valeurs numériques
    good_rows = 0
    for row in table[2:]:
        cells = [str(c).strip() if c else '' for c in row]
        # Label = cellule non numérique non vide
        has_label = any(c for c in cells if c and _parse_fr(c) is None and len(c) > 3)
        # Valeurs dans au moins une autre cellule
        has_nums  = sum(1 for c in cells if _parse_fr(c) is not None) >= 1
        if has_label and has_nums:
            good_rows += 1
    return good_rows >= 5

def _extract_via_tables(page) -> list:
    """Retourne [(label, [val1, val2, ...])] depuis extract_tables()."""
    rows = []
    for t in page.extract_tables():
        if not _table_ok(t): continue
        for row in t:
            if not row: continue
            cells = [str(c).strip().replace('\n', ' ') if c else '' for c in row]
            # Trouver label
            label = next((c for c in cells if c and _parse_fr(c) is None
                          and len(c) > 2 and not SKIP_RE.match(c)), '')
            # Retirer * en préfixe
            label = re.sub(r'^\*\s*', '', label).strip()
            if not label: continue
            vals = [_parse_fr(c) for c in cells if _parse_fr(c) is not None]
            if label and not _should_skip(label):
                rows.append((label, vals))
    return rows

def _is_num_tok(t):
    return bool(re.match(r'^-?\d+$', t.replace(',','').replace('.',''))) \
           and len(t.replace(',','').replace('.','')) >= 1

def _extract_via_xy(page) -> list:
    """Fallback X/Y pour les PDFs sans bordures."""
    words = page.extract_words(x_tolerance=3, y_tolerance=3)
    if not words: return []
    num_words = [w for w in words if _is_num_tok(w['text']) and w['x0'] > 100]
    if not num_words: return []
    thresh = min(w['x0'] for w in num_words) - 5

    lines = defaultdict(list)
    for w in words:
        lines[round(w['top']/3)*3].append(w)

    rows = []
    for y in sorted(lines.keys()):
        row = sorted(lines[y], key=lambda w: w['x0'])
        lw  = [w for w in row if w['x0'] < thresh]
        nw  = [w for w in row if w['x0'] >= thresh and _is_num_tok(w['text'])]

        # Label : filtrer lettres rotatives
        filtered = [w for w in lw
                    if not (len(w['text']) <= 1
                            and re.match(r'^[A-Z.]$', w['text'])
                            and w['x0'] < 50)]
        label = ''
        if filtered:
            label = filtered[0]['text']
            for i in range(1, len(filtered)):
                gap = filtered[i]['x0'] - filtered[i-1]['x1']
                label += filtered[i]['text'] if gap <= 1 else ' ' + filtered[i]['text']
            label = re.sub(r'\s+', ' ', label).strip()
            label = re.sub(r'^\*\s*', '', label).strip()

        # Valeurs fusionnées
        vals = []
        if nw:
            nw_s = sorted(nw, key=lambda w: w['x0'])
            grp  = [nw_s[0]]
            for w in nw_s[1:]:
                if w['x0'] - grp[-1]['x1'] < 18:
                    grp.append(w)
                else:
                    v = _parse_fr(''.join(x['text'] for x in grp))
                    if v is not None: vals.append(v)
                    grp = [w]
            v = _parse_fr(''.join(x['text'] for x in grp))
            if v is not None: vals.append(v)

        if label and not _should_skip(label):
            if vals:
                rows.append((label, vals))
            else:
                prev_label = label  # stocker même sans valeurs
        elif not label and vals and prev_label:
            # Ligne orpheline → rattacher au label précédent
            if rows and rows[-1][0] == prev_label:
                # Fusionner avec la ligne existante
                rows[-1] = (prev_label, rows[-1][1] + vals)
            else:
                rows.append((prev_label, vals))

    return rows

def _extract_page(page):
    rows = _extract_via_tables(page)
    if len(rows) >= 4:
        return rows, 'tables'
    rows = _extract_via_xy(page)
    return rows, 'xy'


# ── Infos générales ───────────────────────────────────────────────────────────
def _extract_info(pdf) -> dict:
    info = {}
    for i in range(min(2, len(pdf.pages))):
        text = pdf.pages[i].extract_text() or ''
        for key, pat in [
            ('raison_sociale',      r'[Rr]aison\s+[Ss]ociale\s*[:\-]?\s*([A-Z][^\n]{3,60})'),
            ('identifiant_fiscal',  r'[Ii]dentifiant\s+[Ff]iscal\s*[:\-]?\s*(\d+)'),
            ('taxe_professionnelle',r'[Tt]axe\s+[Pp]rof\w*\.?\s*[:\-]?\s*([\d\s]+)'),
            ('adresse',             r'[Aa]dresse\s*[:\-]?\s*([^\n]{5,60})'),
        ]:
            if key not in info:
                m = re.search(pat, text, re.IGNORECASE)
                if m: info[key] = m.group(1).strip()

        if 'exercice' not in info:
            for pat in [
                r'p[eé]riode\s+du\s+(\d{2}/\d{2}/\d{4})\s+au\s+(\d{2}/\d{2}/\d{4})',
                r'(\d{2}/\d{2}/\d{4})\s+au\s+(\d{2}/\d{2}/\d{4})',
            ]:
                m = re.search(pat, text, re.I)
                if m:
                    info['exercice']     = f"Du {m.group(1)} au {m.group(2)}"
                    info['exercice_fin'] = m.group(2)
                    break

    for k in ('raison_sociale','identifiant_fiscal','taxe_professionnelle',
              'adresse','exercice','exercice_fin'):
        info.setdefault(k, '')
    return info


# ── Styles Excel ──────────────────────────────────────────────────────────────
def _border():
    s = Side(style='thin', color=C_BORDER)
    return Border(top=s, bottom=s, left=s, right=s)

def _cell(ws, r, c, value=None, bg=C_WHITE, fg='222222', bold=False,
          align='left', num_fmt=None, size=9, wrap=True, indent=0):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font      = Font(name='Arial', size=size, bold=bold, color=fg)
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center',
                               wrap_text=wrap, indent=indent)
    cell.border    = _border()
    if num_fmt: cell.number_format = num_fmt
    return cell

def _row_style(rtype):
    if rtype == 'total':   return C_DARK,  C_WHITE, True
    if rtype == 'result':  return "2E4057", C_WHITE, True
    if rtype == 'section': return C_LIGHT, C_DARK,  True
    return C_WHITE, '333333', False


# ── Construction des feuilles ─────────────────────────────────────────────────
def _write_identification(wb, info):
    ws = wb.create_sheet('1 — Identification')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 50

    ws.merge_cells('A1:B1')
    _cell(ws, 1, 1, 'PIÈCES ANNEXES À LA DÉCLARATION FISCALE',
          bg=C_DARK, fg=C_WHITE, bold=True, align='center', size=13)
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:B2')
    _cell(ws, 2, 1, 'IMPÔTS SUR LES SOCIÉTÉS — Modèle Comptable Normal (loi 9-88)',
          bg=C_MED, fg=C_WHITE, align='center', size=10)
    ws.row_dimensions[2].height = 18

    fields = [
        ('Raison sociale',        info.get('raison_sociale', '—')),
        ('Identifiant fiscal',    info.get('identifiant_fiscal', '—')),
        ('Taxe professionnelle',  info.get('taxe_professionnelle', '—')),
        ('Adresse',               info.get('adresse', '—')),
        ('Exercice',              info.get('exercice', '—')),
    ]
    for i, (lbl, val) in enumerate(fields, 4):
        ws.row_dimensions[i].height = 18
        _cell(ws, i, 1, lbl, bg=C_LIGHT, fg=C_DARK, bold=True, size=9, indent=1)
        _cell(ws, i, 2, val, bg=C_WHITE, fg='222222', size=9, indent=1)


def _write_section(wb, sheet_name, info, headers, rows):
    """
    Écrit une feuille de données.
    headers  : [(col, label_header)]
    rows     : [(label, [val1, val2, ...])]
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    raison   = info.get('raison_sociale', '')
    if_num   = info.get('identifiant_fiscal', '')
    exercice = info.get('exercice', '')
    n_cols   = max(len(headers), max((len(v) for _, v in rows if v), default=0)) + 1

    # Largeurs colonnes
    ws.column_dimensions['A'].width = 46
    for ci in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    # Titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    _cell(ws, 1, 1, sheet_name.split('—')[-1].strip(),
          bg=C_DARK, fg=C_WHITE, bold=True, align='center', size=12)
    ws.row_dimensions[1].height = 22

    # Sous-titre
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols - 1)
    _cell(ws, 2, 1, f'Raison sociale : {raison}',
          bg=C_LIGHT, fg=C_DARK, bold=True, size=9, indent=1)
    _cell(ws, 2, n_cols, f'IF : {if_num}',
          bg=C_LIGHT, fg=C_DARK, align='right', size=9)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
    _cell(ws, 3, 1, f'Exercice : {exercice}',
          bg=C_GRAY, fg='555555', size=9, indent=1)

    ws.row_dimensions[4].height = 4  # séparateur

    # Headers colonnes
    _cell(ws, 5, 1, 'DÉSIGNATION',
          bg=C_MED, fg=C_WHITE, bold=True, align='center', size=9)
    for col_idx, h_label in headers:
        _cell(ws, 5, col_idx, h_label,
              bg=C_MED, fg=C_WHITE, bold=True, align='center', size=9, wrap=True)
    ws.row_dimensions[5].height = 28
    ws.freeze_panes = 'A6'

    # Données
    r = 6
    for label, vals in rows:
        rtype = _row_type(label)
        bg, fg, bold = _row_style(rtype)
        ws.row_dimensions[r].height = 15 if rtype == 'normal' else 17
        indent = 1 if rtype == 'normal' else 0

        _cell(ws, r, 1, label, bg=bg, fg=fg, bold=bold,
              align='left', indent=indent)
        for ci, v in enumerate(vals, 2):
            num = 0 if v is None else v
            _cell(ws, r, ci, num, bg=bg, fg=fg, bold=bold,
                  align='right', num_fmt=NUM_FMT)
        # Cellules vides restantes
        for ci in range(len(vals) + 2, n_cols + 1):
            _cell(ws, r, ci, 0, bg=bg, fg=fg, bold=bold,
                  align='right', num_fmt=NUM_FMT)
        r += 1

    return r - 6


# ── Page mapping ──────────────────────────────────────────────────────────────
AMMC_PAGES = {
    1: ('2 — Bilan Actif',  ['BRUT', 'AMORT. & PROV.', 'NET (N)', 'NET (N-1)']),
    2: ('3 — Bilan Passif', ['EXERCICE N', 'EXERCICE N-1']),
    3: ('4 — CPC',          ["PROPRES À L'EXERCICE", 'EXERCICES PRÉCÉDENTS', 'TOTAUX N', 'TOTAUX N-1']),
    4: ('4 — CPC',          None),   # fusion avec page 3 → même feuille
}

DGI_PAGES = {
    1: ('2 — Bilan Actif',  ['BRUT', 'AMORT. & PROV.', 'NET (N)', 'NET (N-1)']),
    2: ('2 — Bilan Actif',  None),   # fusion avec page 1
    3: ('3 — Bilan Passif', ['EXERCICE N', 'EXERCICE N-1']),
    4: ('4 — CPC',          ["PROPRES À L'EXERCICE", 'EXERCICES PRÉCÉDENTS', 'TOTAUX N', 'TOTAUX N-1']),
    5: ('4 — CPC',          None),
    6: ('4 — CPC',          None),
}


# ── Point d'entrée ────────────────────────────────────────────────────────────
def convert(pdf_path: str, output_path: str) -> dict:
    """Convertit un PDF fiscal en Excel structuré."""
    pdf    = pdfplumber.open(pdf_path)
    n      = len(pdf.pages)
    info   = _extract_info(pdf)
    is_dgi = (n == 7)

    page_map = DGI_PAGES if is_dgi else AMMC_PAGES

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_identification(wb, info)

    # Accumuler les lignes par feuille
    sheets_data = {}  # sheet_name → (headers, rows[])

    for page_idx_0based, page in enumerate(pdf.pages):
        page_idx = page_idx_0based  # 0-based
        # Sauter page 0 (identification)
        if page_idx == 0:
            continue

        cfg = page_map.get(page_idx)
        if cfg is None:
            continue

        sheet_name, headers = cfg

        logger.info(f"Page {page_idx+1} → '{sheet_name}'")
        rows, method = _extract_page(page)
        logger.info(f"  {len(rows)} lignes ({method})")

        if sheet_name not in sheets_data:
            sheets_data[sheet_name] = (headers, [])
        sheets_data[sheet_name][1].extend(rows)

    pdf.close()

    total_rows = 0
    for sheet_name, (headers, rows) in sheets_data.items():
        if not rows:
            continue
        h = [(i+2, lbl) for i, lbl in enumerate(headers)] if headers else []
        n_written = _write_section(wb, sheet_name, info, h, rows)
        total_rows += n_written
        logger.info(f"'{sheet_name}' : {n_written} lignes écrites")

    wb.save(output_path)
    return {
        'info':   info,
        'tables': len(sheets_data),
        'rows':   total_rows,
        'pages':  n,
    }
