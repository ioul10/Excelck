"""
core/ammc_writer.py
Génère un Excel conforme au format AMMC (5 pages) à partir des données extraites.
Structure identique à l'exemple fourni.
"""

import re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.logger import get_logger

logger = get_logger(__name__)

# ── Couleurs ──────────────────────────────────────────────────────────────────
C_DARK   = "1F3864"
C_MED    = "2E75B6"
C_LIGHT  = "D6E4F0"
C_WHITE  = "FFFFFF"
C_GRAY   = "F5F7FA"
C_BORDER = "B8CCE4"
C_TOTAL  = "1F3864"
C_RESULT = "2E4057"
C_SECTION= "D6E4F0"
NUM_FMT  = '#,##0.00;[Red]-#,##0.00;"-"'

def _num(v):
    """Retourne 0 si None, sinon la valeur. Toutes les cellules numériques affichent '-' pour 0."""
    return 0 if v is None else v

# ── Numérotation CPC ──────────────────────────────────────────────────────────
CPC_NUMS = {
    'produits d exploitation':     'I',
    'produits dexploitation':      'I',
    'charges d exploitation':      'II',
    'charges dexploitation':       'II',
    'resultat d exploitation':     'III',
    'produits financiers':         'IV',
    'charges financieres':         'V',
    'charges financières':         'V',
    'resultat financier':          'VI',
    'resultat courant':            'VII',
    'produits non courants':       'VIII',
    'charges non courants':        'IX',
    'charges non courantes':       'IX',
    'resultat non courant':        'X',
    'resultat avant impots':       'XI',
    'impots sur les benefices':    'XII',
    'impots sur les résultats':    'XII',
    'resultat net xi':             'XIII',
    'total des produits':          'XIV',
    'total des charges':           'XV',
    'resultat net total':          'XVI',
}


def _norm(s):
    import unicodedata
    s = unicodedata.normalize('NFD', s).encode('ascii','ignore').decode().lower()
    return re.sub(r'\s+', ' ', re.sub(r'[^\w\s]', ' ', s)).strip()


def _get_cpc_num(label):
    """Retourne le numéro romain pour un label CPC, ou ''."""
    n = _norm(label)
    for key, num in CPC_NUMS.items():
        if key in n:
            return num
    return ''


# ── Styles ────────────────────────────────────────────────────────────────────
def _border():
    s = Side(style='thin', color=C_BORDER)
    return Border(top=s, bottom=s, left=s, right=s)

def _cell(ws, r, c, value=None, bg=C_WHITE, fg='222222', bold=False,
          align='left', num_fmt=None, size=9, wrap=True, indent=0):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font      = Font(name='Arial', size=size, bold=bold, color=fg)
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center',
                               wrap_text=wrap, indent=indent)
    cell.border    = _border()
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def _row_type(label):
    n = _norm(label)
    if re.match(r'^total', n): return 'total'
    if re.match(r'^(resultat|impots)', n): return 'result'
    if label.strip() == label.strip().upper() and len(label.strip()) > 4: return 'section'
    return 'normal'

def _style(row_type):
    if row_type == 'total':   return C_TOTAL,   C_WHITE, True
    if row_type == 'result':  return C_RESULT,  C_WHITE, True
    if row_type == 'section': return C_SECTION, C_DARK,  True
    return C_WHITE, '333333', False


# ── Feuille Identification ────────────────────────────────────────────────────
def _write_identification(wb, info):
    ws = wb.create_sheet('1 — Identification')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 50

    ws.merge_cells('A1:B1')
    _cell(ws, 1, 1, 'PIÈCES ANNEXES À LA DÉCLARATION FISCALE',
          bg=C_DARK, fg=C_WHITE, bold=True, align='center', size=13)
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:B2')
    _cell(ws, 2, 1, 'IMPÔTS SUR LES SOCIÉTÉS — Modèle Comptable Normal (loi 9-88)',
          bg=C_MED, fg=C_WHITE, align='center', size=10)
    ws.row_dimensions[2].height = 18

    fields = [
        ('Raison sociale',       info.get('raison_sociale', '—')),
        ('Identifiant fiscal',   info.get('identifiant_fiscal', '—')),
        ('Taxe professionnelle', info.get('taxe_professionnelle', '—')),
        ('Adresse',              info.get('adresse', '—')),
        ('Exercice',             info.get('exercice', '—')),
    ]
    for i, (lbl, val) in enumerate(fields, 4):
        ws.row_dimensions[i].height = 18
        _cell(ws, i, 1, lbl, bg=C_SECTION, fg=C_DARK, bold=True, size=9, indent=1)
        _cell(ws, i, 2, val, bg=C_WHITE,   fg='222222', size=9, indent=1)


# ── Feuille Bilan Actif ───────────────────────────────────────────────────────
# Structure : A=num, B=label, C=Brut, D=Amort, E=Net N, F=Net N-1
def _write_actif(wb, info, actif_values):
    ws = wb.create_sheet('2 — Bilan Actif')
    ws.sheet_view.showGridLines = False

    # Largeurs
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 46
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    raison   = info.get('raison_sociale', '')
    if_num   = info.get('identifiant_fiscal', '')
    exercice = info.get('exercice', '')

    # Titre + infos
    ws.merge_cells('A1:F1')
    _cell(ws, 1, 1, 'BILAN ACTIF — Modèle Comptable Normal',
          bg=C_DARK, fg=C_WHITE, bold=True, align='center', size=12)
    ws.row_dimensions[1].height = 22

    ws.merge_cells('A2:D2')
    _cell(ws, 2, 1, f'Raison sociale : {raison}',
          bg=C_SECTION, fg=C_DARK, bold=True, size=9, indent=1)
    ws.merge_cells('E2:F2')
    _cell(ws, 2, 5, f'IF : {if_num}',
          bg=C_SECTION, fg=C_DARK, align='right', size=9)

    ws.merge_cells('A3:F3')
    _cell(ws, 3, 1, f'Exercice : {exercice}',
          bg=C_GRAY, fg='555555', size=9, indent=1)

    ws.row_dimensions[4].height = 4  # séparateur

    # Headers colonnes
    _cell(ws, 5, 1, 'ACTIF',          bg=C_MED, fg=C_WHITE, bold=True, align='center', size=9)
    _cell(ws, 5, 2, '',               bg=C_MED, fg=C_WHITE)
    _cell(ws, 5, 3, 'BRUT',           bg=C_MED, fg=C_WHITE, bold=True, align='center', size=9)
    _cell(ws, 5, 4, 'AMORT. & PROV.', bg=C_MED, fg=C_WHITE, bold=True, align='center', size=9, wrap=True)
    _cell(ws, 5, 5, 'NET — EXERCICE N',   bg=C_DARK, fg=C_WHITE, bold=True, align='center', size=9, wrap=True)
    _cell(ws, 5, 6, 'NET — EXERCICE N-1', bg=C_MED,  fg=C_WHITE, bold=True, align='center', size=9, wrap=True)
    ws.row_dimensions[5].height = 28
    ws.freeze_panes = 'B6'

    r = 6
    for label, vals in actif_values.items():
        rtype = _row_type(label)
        bg, fg, bold = _style(rtype)
        ws.row_dimensions[r].height = 15 if rtype == 'normal' else 17

        brut  = vals[0] if len(vals) > 0 else None
        amort = vals[1] if len(vals) > 1 else None
        net_n1= vals[2] if len(vals) > 2 else None
        # Net N = Brut - Amort si les deux sont connus
        net_n = None
        if brut is not None and amort is not None:
            net_n = round(brut - amort, 2)
        elif brut is not None and amort is None:
            net_n = brut

        _cell(ws, r, 1, '',    bg=bg, fg=fg)
        _cell(ws, r, 2, label, bg=bg, fg=fg, bold=bold, indent=1 if rtype=='normal' else 0)
        _cell(ws, r, 3, _num(brut),  bg=bg, fg=fg, bold=bold, align='right', num_fmt=NUM_FMT)
        _cell(ws, r, 4, _num(amort), bg=bg, fg=fg, bold=bold, align='right', num_fmt=NUM_FMT)
        _cell(ws, r, 5, _num(net_n),  bg=bg, fg=fg, bold=bold, align='right', num_fmt=NUM_FMT)
        _cell(ws, r, 6, _num(net_n1), bg=bg, fg=fg, bold=bold, align='right', num_fmt=NUM_FMT)
        r += 1

    return r - 6


# ── Feuille Bilan Passif ──────────────────────────────────────────────────────
# Structure : A=num, B=label, C=Exercice N, D=Exercice N-1
def _write_passif(wb, info, passif_values):
    ws = wb.create_sheet('3 — Bilan Passif')
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 48
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    raison   = info.get('raison_sociale', '')
    if_num   = info.get('identifiant_fiscal', '')
    exercice = info.get('exercice', '')

    ws.merge_cells('A1:D1')
    _cell(ws, 1, 1, 'BILAN PASSIF — Modèle Comptable Normal',
          bg=C_DARK, fg=C_WHITE, bold=True, align='center', size=12)
    ws.row_dimensions[1].height = 22

    ws.merge_cells('A2:B2')
    _cell(ws, 2, 1, f'Raison sociale : {raison}',
          bg=C_SECTION, fg=C_DARK, bold=True, size=9, indent=1)
    ws.merge_cells('C2:D2')
    _cell(ws, 2, 3, f'IF : {if_num}',
          bg=C_SECTION, fg=C_DARK, align='right', size=9)

    ws.merge_cells('A3:D3')
    _cell(ws, 3, 1, f'Exercice : {exercice}',
          bg=C_GRAY, fg='555555', size=9, indent=1)

    ws.row_dimensions[4].height = 4

    _cell(ws, 5, 1, 'PASSIF',        bg=C_MED, fg=C_WHITE, bold=True, align='center', size=9)
    _cell(ws, 5, 2, '',              bg=C_MED, fg=C_WHITE)
    _cell(ws, 5, 3, 'EXERCICE N',    bg=C_DARK, fg=C_WHITE, bold=True, align='center', size=9)
    _cell(ws, 5, 4, 'EXERCICE N-1',  bg=C_MED,  fg=C_WHITE, bold=True, align='center', size=9)
    ws.row_dimensions[5].height = 22
    ws.freeze_panes = 'B6'

    r = 6
    for label, vals in passif_values.items():
        rtype = _row_type(label)
        bg, fg, bold = _style(rtype)
        ws.row_dimensions[r].height = 15 if rtype == 'normal' else 17

        val_n  = vals[0] if len(vals) > 0 else None
        val_n1 = vals[1] if len(vals) > 1 else None

        _cell(ws, r, 1, '',     bg=bg, fg=fg)
        _cell(ws, r, 2, label,  bg=bg, fg=fg, bold=bold, indent=1 if rtype=='normal' else 0)
        _cell(ws, r, 3, _num(val_n),  bg=bg, fg=fg, bold=bold, align='right', num_fmt=NUM_FMT)
        _cell(ws, r, 4, _num(val_n1), bg=bg, fg=fg, bold=bold, align='right', num_fmt=NUM_FMT)
        r += 1

    # Note légale
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(row=r, column=1,
                value='(1) Capital personnel débiteur.  (2) Bénéficiaire (+) / Déficitaire (−).')
    c.font = Font(name='Arial', italic=True, size=8, color='888888')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    return r - 6


# ── Feuille CPC ───────────────────────────────────────────────────────────────
# Structure : A=num romain, B=label, C=Propres N, D=Exerc.Préc, E=Total N, F=Total N-1
def _write_cpc(wb, info, cpc_values):
    ws = wb.create_sheet('4 — CPC')
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 46
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    raison   = info.get('raison_sociale', '')
    if_num   = info.get('identifiant_fiscal', '')
    exercice = info.get('exercice', '')

    ws.merge_cells('A1:F1')
    _cell(ws, 1, 1, 'COMPTE DE PRODUITS ET CHARGES (Hors Taxes)',
          bg=C_DARK, fg=C_WHITE, bold=True, align='center', size=12)
    ws.row_dimensions[1].height = 22

    ws.merge_cells('A2:D2')
    _cell(ws, 2, 1, f'Raison sociale : {raison}',
          bg=C_SECTION, fg=C_DARK, bold=True, size=9, indent=1)
    ws.merge_cells('E2:F2')
    _cell(ws, 2, 5, f'IF : {if_num}',
          bg=C_SECTION, fg=C_DARK, align='right', size=9)

    ws.merge_cells('A3:F3')
    _cell(ws, 3, 1, f'Exercice : {exercice}',
          bg=C_GRAY, fg='555555', size=9, indent=1)

    ws.row_dimensions[4].height = 4

    _cell(ws, 5, 1, '',                 bg=C_MED, fg=C_WHITE, bold=True, align='center', size=9)
    _cell(ws, 5, 2, 'DÉSIGNATION',      bg=C_MED, fg=C_WHITE, bold=True, align='center', size=9)
    _cell(ws, 5, 3, "PROPRES À\nL'EXERCICE",      bg=C_MED, fg=C_WHITE, bold=True, align='center', size=9, wrap=True)
    _cell(ws, 5, 4, 'EXERCICES\nPRÉCÉDENTS',       bg=C_MED, fg=C_WHITE, bold=True, align='center', size=9, wrap=True)
    _cell(ws, 5, 5, 'TOTAUX\nEXERCICE N',           bg=C_DARK,fg=C_WHITE, bold=True, align='center', size=9, wrap=True)
    _cell(ws, 5, 6, 'TOTAUX\nEXERCICE N-1',         bg=C_MED, fg=C_WHITE, bold=True, align='center', size=9, wrap=True)
    ws.row_dimensions[5].height = 30
    ws.freeze_panes = 'B6'

    r = 6
    for label, vals in cpc_values.items():
        rtype  = _row_type(label)
        bg, fg, bold = _style(rtype)
        ws.row_dimensions[r].height = 15 if rtype == 'normal' else 17

        propre_n = vals[0] if len(vals) > 0 else None
        prec_n   = vals[1] if len(vals) > 1 else None
        total_n1 = vals[2] if len(vals) > 2 else None
        # Total N = propre + prec
        total_n = None
        if propre_n is not None and prec_n is not None:
            total_n = round(propre_n + prec_n, 2)
        elif propre_n is not None:
            total_n = propre_n

        num = _get_cpc_num(label)

        _cell(ws, r, 1, num,      bg=bg, fg=fg if rtype not in ('total','result') else C_WHITE,
              bold=True, align='center', size=8)
        _cell(ws, r, 2, label,    bg=bg, fg=fg, bold=bold, indent=1 if rtype=='normal' else 0)
        _cell(ws, r, 3, _num(propre_n), bg=bg, fg=fg, bold=bold, align='right', num_fmt=NUM_FMT)
        _cell(ws, r, 4, _num(prec_n),   bg=bg, fg=fg, bold=bold, align='right', num_fmt=NUM_FMT)
        _cell(ws, r, 5, _num(total_n),  bg=bg, fg=fg, bold=bold, align='right', num_fmt=NUM_FMT)
        _cell(ws, r, 6, _num(total_n1), bg=bg, fg=fg, bold=bold, align='right', num_fmt=NUM_FMT)
        r += 1

    # Notes légales
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(row=r, column=1,
                value='(1) Stock final − Stock initial.   (2) Achats revendus ou consommés = Achats − variation de stock.')
    c.font = Font(name='Arial', italic=True, size=8, color='888888')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    return r - 6


# ── Point d'entrée ────────────────────────────────────────────────────────────
def write(extracted: dict, output_path: str) -> dict:
    """
    Génère l'Excel AMMC depuis les données extraites par pdf_parser.
    extracted = {'info': {...}, 'actif_values': {...}, 'passif_values': {...}, 'cpc_values': {...}}
    """
    info   = extracted.get('info', {})
    actif  = extracted.get('actif_values', {})
    passif = extracted.get('passif_values', {})
    cpc    = extracted.get('cpc_values', {})

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_identification(wb, info)
    n_actif  = _write_actif(wb, info, actif)
    n_passif = _write_passif(wb, info, passif)
    n_cpc    = _write_cpc(wb, info, cpc)

    wb.save(output_path)

    logger.info(f"AMMC Excel: {n_actif} actif + {n_passif} passif + {n_cpc} CPC lignes")
    return {
        'actif':  n_actif,
        'passif': n_passif,
        'cpc':    n_cpc,
        'total':  n_actif + n_passif + n_cpc,
    }
