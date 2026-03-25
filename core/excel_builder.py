"""
core/excel_builder.py
Étape 4 : Génération du classeur Excel professionnel
         - Formules dynamiques (NET = Brut - Amort, Totaux, Résultats)
         - Liens inter-feuilles pour le tableau de bord
         - Mise en forme colorée aux standards financiers
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.fiscal_schema import (
    SECTION_LABELS, TOTAL_LABELS, SUBTOTAL_LABELS, RESULT_LABELS,
    ACTIF_FORMULAS, PASSIF_FORMULAS, CPC_FORMULAS, CPC_DIFFERENCES, CPC_ADDITIONS,
)
from utils.logger import get_logger

logger = get_logger(__name__)

# ── Palette couleurs ──────────────────────────────────────────────────────────
C = {
    "dark_blue":   "1F3864",
    "mid_blue":    "2E75B6",
    "light_blue":  "BDD7EE",
    "very_light":  "DEEAF1",
    "white":       "FFFFFF",
    "total_bg":    "1F3864",
    "subtotal_bg": "2E75B6",
    "gold":        "C9A84C",
    "green_bg":    "E2EFDA",
    "red_light":   "FCE4D6",
    "gray":        "F2F2F2",
    "section_bg":  "1F3864",
}

NUM_FMT = '#,##0.00;(#,##0.00);"-"'


def _border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _font(bold=False, color="000000", size=10, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)


def _align(h="left", v="center", indent=0, wrap=False):
    return Alignment(horizontal=h, vertical=v, indent=indent, wrap_text=wrap)


class ExcelBuilder:

    def __init__(self, fiscal_data: dict,
                 with_dashboard=True,
                 with_formulas=True,
                 with_colors=True):
        self.data          = fiscal_data
        self.with_dashboard = with_dashboard
        self.with_formulas  = with_formulas
        self.with_colors    = with_colors
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

        # Carte des lignes pour formules inter-lignes
        self._row_maps = {}   # sheet_name -> {label: row_number}

    def build(self, output_path: str) -> dict:
        info   = self.data.get("info", {})
        actif  = self.data.get("bilan_actif", [])
        passif = self.data.get("bilan_passif", [])
        cpc    = self.data.get("cpc", [])

        self._build_info(info)
        self._build_bilan_actif(actif, info)
        self._build_bilan_passif(passif, info)
        self._build_cpc(cpc, info)

        if self.with_dashboard:
            self._build_dashboard(info)

        self.wb.save(output_path)

        stats = {
            "sheets":   len(self.wb.sheetnames),
            "formulas": self._count_formulas(),
            "rows":     len(actif) + len(passif) + len(cpc),
        }
        logger.info(f"Excel généré : {stats}")
        return stats

    # ── FEUILLE 1 : Infos générales ───────────────────────────────────────────

    def _build_info(self, info: dict):
        ws = self.wb.create_sheet("1 - Infos Générales")
        ws.sheet_view.showGridLines = False

        self._title_row(ws, 1, 2,
            "PIÈCES ANNEXES À LA DÉCLARATION FISCALE", C["dark_blue"], size=14)
        self._title_row(ws, 2, 2,
            "IMPÔTS SUR LES SOCIÉTÉS — Modèle Comptable Normal", C["mid_blue"], size=11)
        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 24

        headers = [("Champ", C["mid_blue"]), ("Valeur", C["mid_blue"])]
        for ci, (h, bg) in enumerate(headers, 1):
            c = ws.cell(row=3, column=ci, value=h)
            self._header_cell(c, bg=bg)
        ws.row_dimensions[3].height = 24

        display_keys = [
            ("Raison sociale",        "raison_sociale"),
            ("Taxe professionnelle",  "taxe_professionnelle"),
            ("Identifiant fiscal",    "identifiant_fiscal"),
            ("Adresse",               "adresse"),
            ("Exercice",              "exercice"),
            ("Date de déclaration",   "date_declaration"),
        ]
        for r, (display, key) in enumerate(display_keys, 4):
            bg = C["very_light"] if r % 2 == 0 else C["white"]
            kc = ws.cell(row=r, column=1, value=display)
            vc = ws.cell(row=r, column=2, value=info.get(key, "—"))
            for c in [kc, vc]:
                c.fill   = _fill(bg)
                c.font   = _font(bold=(c == kc), color=C["dark_blue"])
                c.border = _border()
                c.alignment = _align("left", indent=1)
            ws.row_dimensions[r].height = 22

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60

    # ── FEUILLE 2 : Bilan Actif ───────────────────────────────────────────────

    def _build_bilan_actif(self, rows: list, info: dict):
        ws = self.wb.create_sheet("2 - Bilan Actif")
        ws.sheet_view.showGridLines = False
        sheet_name = "2 - Bilan Actif"

        rs = info.get("raison_sociale", "")
        self._title_row(ws, 1, 5, f"BILAN — ACTIF  |  {info.get('exercice','')}", C["dark_blue"], size=13)
        self._title_row(ws, 2, 5, f"{rs}  —  IF: {info.get('identifiant_fiscal','')}", C["mid_blue"], size=10)
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 22

        for ci, h in enumerate(["ACTIF", "BRUT (N)", "AMORT. & PROV.", "NET (N)", "NET (N-1)"], 1):
            self._header_cell(ws.cell(row=3, column=ci, value=h))
        ws.row_dimensions[3].height = 28
        ws.freeze_panes = "A4"

        row_map = {}
        r = 4
        for item in rows:
            label = item[0] if item else ""
            vals  = list(item[1:]) + [None] * 5

            ws.row_dimensions[r].height = 18
            row_map[label] = r

            kind = self._row_kind(label)

            if kind == "section":
                self._section_row(ws, r, f"  ▶  {label}", 5)
                r += 1
                continue

            # Libellé
            indent = 2 if label.startswith("  ") else 0
            lc = ws.cell(row=r, column=1, value=label.strip())
            self._style_data_cell(lc, kind, r, align="left", indent=indent)

            # Brut, Amort
            for ci, vi in enumerate([0, 1], 2):
                c = ws.cell(row=r, column=ci)
                if vals[vi] is not None:
                    c.value = vals[vi]
                self._style_data_cell(c, kind, r, align="right")

            # Net N : formule = Brut - Amort si pas total/section
            net_cell = ws.cell(row=r, column=4)
            if self.with_formulas and kind == "data" and vals[0] is not None:
                net_cell.value = f"=B{r}-C{r}"
            elif vals[2] is not None:
                net_cell.value = vals[2]
            self._style_data_cell(net_cell, kind, r, align="right", is_formula=True)

            # Net N-1
            n1_cell = ws.cell(row=r, column=5)
            if vals[3] is not None:
                n1_cell.value = vals[3]
            self._style_data_cell(n1_cell, kind, r, align="right")

            r += 1

        # Appliquer formules de totaux
        if self.with_formulas:
            self._apply_sum_formulas(ws, row_map, ACTIF_FORMULAS, cols=[2, 3, 4, 5])
            # NET des sous-totaux = Brut - Amort
            for lbl in SUBTOTAL_LABELS:
                if lbl in row_map:
                    rr = row_map[lbl]
                    ws.cell(row=rr, column=4).value = f"=B{rr}-C{rr}"

        self._row_maps[sheet_name] = row_map
        self._set_col_widths(ws, {"A": 50, "B": 18, "C": 18, "D": 18, "E": 18})

    # ── FEUILLE 3 : Bilan Passif ──────────────────────────────────────────────

    def _build_bilan_passif(self, rows: list, info: dict):
        ws = self.wb.create_sheet("3 - Bilan Passif")
        ws.sheet_view.showGridLines = False
        sheet_name = "3 - Bilan Passif"

        rs = info.get("raison_sociale", "")
        self._title_row(ws, 1, 3, f"BILAN — PASSIF  |  {info.get('exercice','')}", C["dark_blue"], size=13)
        self._title_row(ws, 2, 3, f"{rs}  —  IF: {info.get('identifiant_fiscal','')}", C["mid_blue"], size=10)
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 22

        for ci, h in enumerate(["PASSIF", "EXERCICE N", "EXERCICE N-1"], 1):
            self._header_cell(ws.cell(row=3, column=ci, value=h))
        ws.row_dimensions[3].height = 28
        ws.freeze_panes = "A4"

        row_map = {}
        r = 4
        for item in rows:
            label = item[0] if item else ""
            vals  = list(item[1:]) + [None, None]

            ws.row_dimensions[r].height = 18
            row_map[label] = r

            kind = self._row_kind(label)
            if kind == "section":
                self._section_row(ws, r, f"  ▶  {label}", 3)
                r += 1
                continue

            indent = 2 if label.startswith("  ") else 0
            lc = ws.cell(row=r, column=1, value=label.strip())
            self._style_data_cell(lc, kind, r, align="left", indent=indent)

            for ci, vi in enumerate([0, 1], 2):
                c = ws.cell(row=r, column=ci)
                if vals[vi] is not None:
                    c.value = vals[vi]
                self._style_data_cell(c, kind, r, align="right")

            r += 1

        if self.with_formulas:
            self._apply_sum_formulas(ws, row_map, PASSIF_FORMULAS, cols=[2, 3])

        self._row_maps[sheet_name] = row_map
        self._set_col_widths(ws, {"A": 54, "B": 22, "C": 22})

    # ── FEUILLE 4 : CPC ───────────────────────────────────────────────────────

    def _build_cpc(self, rows: list, info: dict):
        ws = self.wb.create_sheet("4 - CPC")
        ws.sheet_view.showGridLines = False
        sheet_name = "4 - CPC"

        rs = info.get("raison_sociale", "")
        self._title_row(ws, 1, 5,
            f"COMPTE DE PRODUITS ET CHARGES  |  {info.get('exercice','')}", C["dark_blue"], size=13)
        self._title_row(ws, 2, 5, f"{rs}  —  Hors Taxes", C["mid_blue"], size=10)
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 22

        for ci, h in enumerate([
            "DÉSIGNATION", "Propres à l'exercice N",
            "Exercices précédents", "TOTAL N", "TOTAL N-1"
        ], 1):
            c = ws.cell(row=3, column=ci, value=h)
            self._header_cell(c)
        ws.row_dimensions[3].height = 36
        ws.freeze_panes = "A4"

        row_map = {}
        r = 4
        for item in rows:
            label = item[0] if item else ""
            vals  = list(item[1:]) + [None] * 4

            ws.row_dimensions[r].height = 18
            row_map[label] = r

            kind = self._row_kind(label)
            if kind == "section":
                self._section_row(ws, r, f"  ▶  {label}", 5)
                r += 1
                continue

            indent = 2 if label.startswith("  ") else 0
            lc = ws.cell(row=r, column=1, value=label.strip())
            self._style_data_cell(lc, kind, r, align="left", indent=indent)

            # Propre N, Exercices Précédents, Total N, Total N-1
            for ci, vi in enumerate([0, 1, 2, 3], 2):
                c = ws.cell(row=r, column=ci)
                if vals[vi] is not None:
                    c.value = vals[vi]
                self._style_data_cell(c, kind, r, align="right",
                                      is_formula=(ci == 4))

            # Total N = col B + col C
            if self.with_formulas and kind not in ("section",):
                tot_cell = ws.cell(row=r, column=4)
                if not isinstance(tot_cell.value, str):  # pas déjà une formule
                    tot_cell.value = f"=B{r}+C{r}"

            r += 1

        if self.with_formulas:
            self._apply_sum_formulas(ws, row_map, CPC_FORMULAS, cols=[2, 3, 4, 5])
            self._apply_cpc_differences(ws, row_map)

        self._row_maps[sheet_name] = row_map
        self._set_col_widths(ws, {"A": 54, "B": 20, "C": 20, "D": 20, "E": 20})

    # ── FEUILLE 5 : Tableau de Bord ───────────────────────────────────────────

    def _build_dashboard(self, info: dict):
        ws = self.wb.create_sheet("5 - Tableau de Bord")
        ws.sheet_view.showGridLines = False

        self._title_row(ws, 1, 5,
            "TABLEAU DE BORD — SYNTHÈSE FINANCIÈRE", C["dark_blue"], size=14)
        self._title_row(ws, 2, 5,
            info.get("raison_sociale", ""), C["mid_blue"], size=11)
        ws.row_dimensions[1].height = 38
        ws.row_dimensions[2].height = 26

        for ci, h in enumerate(
            ["INDICATEUR", "EXERCICE N", "EXERCICE N-1", "Δ (N vs N-1)", "Unité"], 1
        ):
            self._header_cell(ws.cell(row=3, column=ci, value=h))
        ws.row_dimensions[3].height = 26
        ws.freeze_panes = "A4"

        # KPIs avec références vers les autres feuilles
        actif_map  = self._row_maps.get("2 - Bilan Actif", {})
        passif_map = self._row_maps.get("3 - Bilan Passif", {})
        cpc_map    = self._row_maps.get("4 - CPC", {})

        def ref(sheet, label, col):
            rm = {"2 - Bilan Actif": actif_map,
                  "3 - Bilan Passif": passif_map,
                  "4 - CPC": cpc_map}.get(sheet, {})
            row = rm.get(label)
            if row:
                return f"='{sheet}'!{get_column_letter(col)}{row}"
            return None

        kpi_groups = [
            ("📈 COMPTE DE RÉSULTATS", [
                ("Chiffre d'affaires",
                 ref("4 - CPC", "  Chiffre d'affaires", 4),
                 ref("4 - CPC", "  Chiffre d'affaires", 5)),
                ("Produits d'exploitation",
                 ref("4 - CPC", "TOTAL I - PRODUITS D'EXPLOITATION", 4),
                 ref("4 - CPC", "TOTAL I - PRODUITS D'EXPLOITATION", 5)),
                ("Charges d'exploitation",
                 ref("4 - CPC", "TOTAL II - CHARGES D'EXPLOITATION", 4),
                 ref("4 - CPC", "TOTAL II - CHARGES D'EXPLOITATION", 5)),
                ("Résultat d'exploitation",
                 ref("4 - CPC", "RÉSULTAT D'EXPLOITATION (I-II)", 4),
                 ref("4 - CPC", "RÉSULTAT D'EXPLOITATION (I-II)", 5)),
                ("Résultat financier",
                 ref("4 - CPC", "RÉSULTAT FINANCIER (IV-V)", 4),
                 ref("4 - CPC", "RÉSULTAT FINANCIER (IV-V)", 5)),
                ("Résultat courant",
                 ref("4 - CPC", "RÉSULTAT COURANT (III+VI)", 4),
                 ref("4 - CPC", "RÉSULTAT COURANT (III+VI)", 5)),
                ("Résultat net de l'exercice",
                 ref("4 - CPC", "RÉSULTAT NET (XI-XII)", 4),
                 ref("4 - CPC", "RÉSULTAT NET (XI-XII)", 5)),
            ]),
            ("📊 BILAN — ACTIF", [
                ("Actif immobilisé net",
                 ref("2 - Bilan Actif", "TOTAL I (A+B+C+D+E)", 4),
                 ref("2 - Bilan Actif", "TOTAL I (A+B+C+D+E)", 5)),
                ("Actif circulant net",
                 ref("2 - Bilan Actif", "TOTAL II (F+G+H+I)", 4),
                 ref("2 - Bilan Actif", "TOTAL II (F+G+H+I)", 5)),
                ("Trésorerie actif",
                 ref("2 - Bilan Actif", "TOTAL III", 4),
                 ref("2 - Bilan Actif", "TOTAL III", 5)),
                ("TOTAL GÉNÉRAL ACTIF",
                 ref("2 - Bilan Actif", "TOTAL GÉNÉRAL (I+II+III)", 4),
                 ref("2 - Bilan Actif", "TOTAL GÉNÉRAL (I+II+III)", 5)),
            ]),
            ("📊 BILAN — PASSIF", [
                ("Capitaux propres",
                 ref("3 - Bilan Passif", "Total des capitaux propres (A)", 2),
                 ref("3 - Bilan Passif", "Total des capitaux propres (A)", 3)),
                ("Subventions d'investissement",
                 ref("3 - Bilan Passif", "  Subventions d'investissement", 2),
                 ref("3 - Bilan Passif", "  Subventions d'investissement", 3)),
                ("Financement permanent",
                 ref("3 - Bilan Passif", "TOTAL I (A+B+C+D+E) PASSIF", 2),
                 ref("3 - Bilan Passif", "TOTAL I (A+B+C+D+E) PASSIF", 3)),
                ("TOTAL GÉNÉRAL PASSIF",
                 ref("3 - Bilan Passif", "TOTAL GÉNÉRAL PASSIF (I+II+III)", 2),
                 ref("3 - Bilan Passif", "TOTAL GÉNÉRAL PASSIF (I+II+III)", 3)),
            ]),
        ]

        r = 4
        for group_label, kpis in kpi_groups:
            # Ligne de groupe
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            gc = ws.cell(row=r, column=1, value=group_label)
            gc.font      = _font(bold=True, color=C["white"], size=11)
            gc.fill      = _fill(C["dark_blue"])
            gc.alignment = _align("left", indent=1)
            gc.border    = _border()
            ws.row_dimensions[r].height = 24
            r += 1

            for label, fn, fn1 in kpis:
                bg = C["very_light"] if r % 2 == 0 else C["white"]
                ws.row_dimensions[r].height = 22

                lc = ws.cell(row=r, column=1, value=label)
                lc.font = _font(size=10)
                lc.fill = _fill(bg)
                lc.alignment = _align("left", indent=1)
                lc.border = _border()

                for ci, val in enumerate([fn, fn1], 2):
                    c = ws.cell(row=r, column=ci)
                    if val:
                        c.value = val
                    c.number_format = NUM_FMT
                    c.font      = _font(size=10, color=C["dark_blue"])
                    c.fill      = _fill(bg)
                    c.alignment = _align("right")
                    c.border    = _border()

                # Delta N vs N-1
                dc = ws.cell(row=r, column=4)
                if fn and fn1:
                    dc.value = f"=B{r}-C{r}"
                    dc.number_format = NUM_FMT
                    dc.font      = _font(size=10, color=C["dark_blue"])
                dc.fill      = _fill(bg)
                dc.alignment = _align("right")
                dc.border    = _border()

                uc = ws.cell(row=r, column=5, value="MAD")
                uc.font      = _font(size=9, color="888888")
                uc.fill      = _fill(bg)
                uc.alignment = _align("center")
                uc.border    = _border()

                r += 1

        self._set_col_widths(ws, {"A": 42, "B": 22, "C": 22, "D": 22, "E": 8})

    # ── Formules CPC : différences/additions ─────────────────────────────────

    def _apply_cpc_differences(self, ws, row_map: dict):
        for result_label, operands in CPC_DIFFERENCES.items():
            if result_label not in row_map:
                continue
            tr = row_map[result_label]
            is_add = result_label in CPC_ADDITIONS

            if len(operands) == 2:
                a_label, b_label = operands
                ar = row_map.get(a_label)
                br = row_map.get(b_label)
                if ar and br:
                    for col in [2, 3, 4, 5]:
                        cl = get_column_letter(col)
                        op = "+" if is_add else "-"
                        ws.cell(row=tr, column=col).value = f"={cl}{ar}{op}{cl}{br}"

    # ── Helpers de formules ───────────────────────────────────────────────────

    def _apply_sum_formulas(self, ws, row_map: dict, formulas: dict, cols: list):
        for total_label, component_labels in formulas.items():
            if total_label not in row_map:
                continue
            tr = row_map[total_label]
            refs_by_col = {}
            for comp in component_labels:
                cr = row_map.get(comp)
                if cr:
                    for col in cols:
                        refs_by_col.setdefault(col, []).append(
                            f"{get_column_letter(col)}{cr}"
                        )
            for col, refs in refs_by_col.items():
                if refs:
                    ws.cell(row=tr, column=col).value = "=" + "+".join(refs)

    # ── Helpers de style ──────────────────────────────────────────────────────

    def _row_kind(self, label: str) -> str:
        if label in SECTION_LABELS:
            return "section"
        if label in TOTAL_LABELS:
            return "total"
        if label in RESULT_LABELS:
            return "result"
        if label in SUBTOTAL_LABELS:
            return "subtotal"
        return "data"

    def _style_data_cell(self, cell, kind: str, row: int,
                         align="right", indent=0, is_formula=False):
        if not self.with_colors:
            cell.border = _border()
            cell.alignment = _align(align, indent=indent)
            cell.number_format = NUM_FMT
            return

        bg_map = {
            "total":    C["total_bg"],
            "result":   C["gold"],
            "subtotal": C["subtotal_bg"],
        }
        fg_map = {
            "total":    C["white"],
            "result":   C["white"],
            "subtotal": C["white"],
            "data":     "000000",
        }

        bg = bg_map.get(kind, C["very_light"] if row % 2 == 0 else C["white"])
        fg = fg_map.get(kind, "000000")
        bold = kind in ("total", "result", "subtotal")

        cell.fill      = _fill(bg)
        cell.font      = _font(bold=bold, color=fg, size=10)
        cell.alignment = _align(align, indent=indent)
        cell.border    = _border()
        if align == "right":
            cell.number_format = NUM_FMT

    def _header_cell(self, cell, bg=None):
        bg = bg or C["mid_blue"]
        cell.font      = _font(bold=True, color=C["white"], size=10)
        cell.fill      = _fill(bg)
        cell.alignment = _align("center", wrap=True)
        cell.border    = _border()

    def _title_row(self, ws, row: int, colspan: int, text: str, bg: str, size=12):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=colspan)
        c = ws.cell(row=row, column=1, value=text)
        c.font      = _font(bold=True, color=C["white"], size=size)
        c.fill      = _fill(bg)
        c.alignment = _align("center")
        c.border    = _border()

    def _section_row(self, ws, row: int, label: str, colspan: int):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=colspan)
        c = ws.cell(row=row, column=1, value=label)
        c.font      = _font(bold=True, color=C["white"], size=10)
        c.fill      = _fill(C["section_bg"])
        c.alignment = _align("left", indent=1)
        c.border    = _border()

    @staticmethod
    def _set_col_widths(ws, widths: dict):
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    def _count_formulas(self) -> int:
        count = 0
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith("="):
                        count += 1
        return count
