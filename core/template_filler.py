"""
core/template_filler.py
Remplit le template Excel fiscal avec les données extraites du PDF
"""
from openpyxl import load_workbook
from utils.logger import get_logger

logger = get_logger(__name__)

class TemplateFiller:
    """
    Remplit le template Excel standard avec les données fiscales
    """
    
    def __init__(self, template_path: str):
        self.template_path = template_path
        self.wb = load_workbook(template_path, data_only=False)
        logger.info(f"Template chargé : {template_path}")
    
    def fill_from_data(self, fiscal_data: dict, output_path: str) -> dict:
        """
        Remplit toutes les feuilles du template
        """
        info = fiscal_data.get("info", {})
        actif = fiscal_data.get("bilan_actif", [])
        passif = fiscal_data.get("bilan_passif", [])
        cpc = fiscal_data.get("cpc", [])
        
        # Feuille 1: Infos Générales
        self._fill_info_sheet(info)
        
        # Feuille 2: Bilan Actif
        self._fill_bilan_actif(actif)
        
        # Feuille 3: Bilan Passif
        self._fill_bilan_passif(passif)
        
        # Feuille 4: CPC
        self._fill_cpc(cpc)
        
        # Sauvegarder
        self.wb.save(output_path)
        logger.info(f"Template rempli sauvegardé : {output_path}")
        
        return {
            "sheets": len(self.wb.sheetnames),
            "rows_filled": len(actif) + len(passif) + len(cpc)
        }
    
    def _fill_info_sheet(self, info: dict):
        """Remplit la feuille Infos Générales"""
        ws = self.wb["1 - Infos Générales"]
        
        mappings = {
            "raison_sociale": "B5",
            "taxe_professionnelle": "B6",
            "identifiant_fiscal": "B7",
            "adresse": "B8",
            "exercice": "B9",
            "date_declaration": "B10",
        }
        
        for key, cell_ref in mappings.items():
            value = info.get(key, "")
            ws[cell_ref] = value if value else ""
            logger.debug(f"Info {key}: {value}")
    
    def _fill_bilan_actif(self, rows: list):
        """Remplit le Bilan Actif par matching de labels"""
        ws = self.wb["2 - Bilan Actif"]
        
        # Créer un dictionnaire label → valeurs
        data_dict = {}
        for row in rows:
            if row and row[0]:
                label = str(row[0]).strip()
                values = row[1:] if len(row) > 1 else []
                data_dict[label] = values
        
        # Parcourir les lignes du template (commence à la ligne 7)
        for row_idx in range(7, ws.max_row + 1):
            label_cell = ws[f"A{row_idx}"].value
            if label_cell:
                label = str(label_cell).strip()
                if label in data_dict:
                    values = data_dict[label]
                    # Colonnes: B=Brut, C=Amort, D=Net N, E=Net N-1
                    for col_idx, val in enumerate(values[:4], 2):
                        if val is not None and val != "":
                            cell_ref = f"{chr(65 + col_idx)}{row_idx}"  # B=66, C=67...
                            ws[cell_ref] = float(val) if isinstance(val, (int, float)) else val
                    logger.debug(f"Actif: {label} → {values[:4]}")
    
    def _fill_bilan_passif(self, rows: list):
        """Remplit le Bilan Passif"""
        ws = self.wb["3 - Bilan Passif"]
        
        data_dict = {}
        for row in rows:
            if row and row[0]:
                label = str(row[0]).strip()
                values = row[1:] if len(row) > 1 else []
                data_dict[label] = values
        
        # Parcourir les lignes (commence à la ligne 7)
        for row_idx in range(7, ws.max_row + 1):
            label_cell = ws[f"A{row_idx}"].value
            if label_cell:
                label = str(label_cell).strip()
                if label in data_dict:
                    values = data_dict[label]
                    # Colonnes: B=Exercice N, C=Exercice N-1
                    for col_idx, val in enumerate(values[:2], 2):
                        if val is not None and val != "":
                            cell_ref = f"{chr(65 + col_idx)}{row_idx}"
                            ws[cell_ref] = float(val) if isinstance(val, (int, float)) else val
                    logger.debug(f"Passif: {label} → {values[:2]}")
    
    def _fill_cpc(self, rows: list):
        """Remplit le CPC"""
        ws = self.wb["4 - CPC"]
        
        data_dict = {}
        for row in rows:
            if row and row[0]:
                label = str(row[0]).strip()
                values = row[1:] if len(row) > 1 else []
                data_dict[label] = values
        
        # Parcourir les lignes (commence à la ligne 7)
        for row_idx in range(7, ws.max_row + 1):
            label_cell = ws[f"A{row_idx}"].value
            if label_cell:
                label = str(label_cell).strip()
                if label in data_dict:
                    values = data_dict[label]
                    # Colonnes: B=Propre N, C=Exerc Préc, D=Total N, E=Total N-1
                    for col_idx, val in enumerate(values[:4], 2):
                        if val is not None and val != "":
                            cell_ref = f"{chr(65 + col_idx)}{row_idx}"
                            ws[cell_ref] = float(val) if isinstance(val, (int, float)) else val
                    logger.debug(f"CPC: {label} → {values[:4]}")
